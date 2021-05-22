'''
Author: what-is-me
E-mail: nt_cqc@126.com
Github: https://github.com/what-is-me
LeetCode: https://leetcode-cn.com/u/what-is-me/
Date: 2021-05-19 14:16:38
LastEditors: what-is-me
LastEditTime: 2021-05-21 22:35:44
Description: 输入单词，查询并写入xlsx或md或txt，或者不写入。
'''
'''
# -*- coding:utf-8 -*-
print("\033[1;30m 字体颜色：白色\033[0m")
print("\033[1;31m 字体颜色：红色\033[0m")
print("\033[1;32m 字体颜色：深黄色\033[0m")
print("\033[1;33m 字体颜色：浅黄色\033[0m")
print("\033[1;34m 字体颜色：蓝色\033[0m")
print("\033[1;35m 字体颜色：淡紫色\033[0m")
print("\033[1;36m 字体颜色：青色\033[0m")
print("\033[1;37m 字体颜色：灰色\033[0m")
print("\033[1;38m 字体颜色：浅灰色\033[0m")
 
print("背景颜色：白色   \033[1;40m    \033[0m")
print("背景颜色：红色   \033[1;41m    \033[0m")
print("背景颜色：深黄色 \033[1;42m    \033[0m")
print("背景颜色：浅黄色 \033[1;43m    \033[0m")
print("背景颜色：蓝色   \033[1;44m    \033[0m")
print("背景颜色：淡紫色 \033[1;45m    \033[0m")
print("背景颜色：青色   \033[1;46m    \033[0m")
print("背景颜色：灰色   \033[1;47m    \033[0m")
'''




import wordlistenquiry as wl
from openpyxl import Workbook
def _yellow(string):
    return "\033[1;33m"+string+"\033[0m"


def _red(string):
    return "\033[1;31m"+string+"\033[0m"


def md(title):
    w = open(title + ".md", "w")
    w.write("# "+title+"\n")
    w.write("|单词|释义|\n")
    w.write("|:-:|:-:|\n")
    w.close()
    s = ""
    while(True):
        s = input(_yellow(">>>"))
        if s == "" or s == "\\end":
            break
        w = open(title + ".md", "a+")
        ex = wl.search(s)
        if ex == "未收录":
            print(_red(ex))
            continue
        print(ex)
        w.write("|"+s+"|"+ex+"|\n")
        w.close()


def text(title):
    w = open(title + ".txt", "w")
    w.write(title.upper()+"\n")
    for i in range(20):
        w.write("==")
    w.write("\n\n")
    w.close()
    s = ""
    while(True):
        s = input(_yellow(">>>"))
        if s == "" or s == "\\end":
            break
        w = open(title + ".txt", "a+")
        ex = wl.search(s)
        if ex == "未收录":
            print(_red(ex))
            continue
        print(ex)
        w.write(s+":\t"+ex+"\n")
        w.close()


def xlsx(title):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    line = 1
    ws.cell(row=line, column=1, value="单词")
    ws.cell(row=line, column=2, value="释义")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 100
    wb.save(title+".xlsx")
    s = ""
    line = 2
    while(True):
        s = input(_yellow(">>>"))
        if s == "" or s == "\\end":
            break
        ex = wl.search(s)
        if ex == "未收录":
            print(_red(ex))
            continue
        print(ex)
        ws.cell(row=line, column=1, value=s)
        ws.cell(row=line, column=2, value=ex)
        wb.save(title+".xlsx")
        line += 1


title = input(_yellow("title?:"))
if title != "":
    ch = input(_yellow("md/xlsx/txt:"))
    if ch == "" or ch == "txt":
        text(title)
    elif ch == "md":
        md(title)
    elif ch == "xlsx":
        xlsx(title)
    else:
        print(_red("error\n"))
else:
    while(True):
        s = input(_yellow(">>>"))
        if s == "" or s == "\\end":
            break
        ex = wl.search(s)
        if ex == "未收录":
            print(_red(ex))
            continue
        print(ex)
