# =======================================================================
#     @Author:         what-is-me
#     @Date:           2021-04-29 19:00:39
#     @LastEditors:    what-is-me
#     @LastEditTime:   2021-04-29 19:00:40
#     @Description:    一个简单的爬虫程序，用于批量查询单词，有图形界面。
#     @Address:        https://github.com/what-is-me/wordlisttranslate/
# =======================================================================


# -*- coding: utf-8 -*-


# =======================================================================
# import:
import codecs
import re
import tkinter.messagebox as messagebox
from tkinter import *
from tkinter.filedialog import *
import openpyxl
import requests
from openpyxl.styles import Color, Fill, Font
# =======================================================================


# =======================================================================
# functions:
def getHtml(url):
    # 获得网址源代码
    headers = {
        "User-Agent": "User-Agent:Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0;"}
    page = requests.get(url, headers=headers)
    page.encoding = 'utf-8'
    html = page.text
    return html


class getimg:
    def youdao(html):
        html = html.split('<div class="trans-container">')[-1]
        html = html.split('<!--网络翻译-->')[0]
        j = 0
        reg = r'<li>(.*?)</li>'
        img = re.compile(reg)
        img_list = re.findall(img, html)
        if len(img_list) == 0:
            result = "未收录"
            return result
        for i in range(0, len(img_list)):
            element = img_list[i]
            finding = '. '
            num = element.find(finding)
            if num >= 0:
                j = i
        result = ''
        for c in range(0, j+1):
            result = result + img_list[c] + "；"
        result = re.sub('<li>', '\t', result)
        result = re.sub(r'<(.*?)>', '\t', result)
        return result

    def jinshan(html):
        j = 0
        reg = r'<ul class="Mean_part__1RA2V"><li>(.*?)</ul>'
        img = re.compile(reg)
        img_list = re.findall(img, html)
        if len(img_list) == 0:
            result = "未收录"
            return result
        for i in range(0, len(img_list)):
            element = img_list[i]
            finding = '. '
            num = element.find(finding)
            if num >= 0:
                j = i
        result = ''
        for c in range(0, j+1):
            result = result + img_list[c]
        result = re.sub('<i>', '\t', result)
        result = re.sub('&lt;', '[', result)
        result = re.sub('&gt;', ']', result)
        return re.sub(r'<(.*?)>', '', result)

    def bing(html):
        j = 0
        reg = r'<meta name="description" content="(.*?)" />'
        img = re.compile(reg)
        img_list = re.search(img, html).group()
        img_list = img_list.split('<meta name="description" content="')[-1]
        img_list = img_list.split('" />')[0]
        if len(img_list) == 0:
            result = "未收录"
            return result
        for i in range(0, len(img_list)):
            element = img_list[i]
            num = len(element)
            if num >= 0:
                j = i
        result = ''
        for c in range(0, j+1):
            result = result + img_list[c]
        result = re.sub('必应词典为您提供', '', result)
        result = re.sub('的释义', '', result)
        result = re.sub('英', '', result)
        result = re.sub('美', '', result)
        result = re.sub('，', '', result)
        result = result.split('网络释义：')[0]
        return re.sub(r'\[(.*?)\]', '', result)

    def haici(html):
        html = html.split('<div class="basic clearfix">')[-1]
        html = html.split('<li style="padding-top: 25px;">')[0]
        reg1 = r'<span>(.*?)</span>'
        img1 = re.compile(reg1)
        img_list1 = re.findall(img1, html)
        reg2 = r'<strong>(.*?)</strong>'
        img2 = re.compile(reg2)
        img_list2 = re.findall(img2, html)
        if len(img_list2) == 0:
            result = "未收录"
            return result
        result = ''
        if(len(img_list1) == 0):
            for i in range(0, len(img_list2)):
                result += img_list2[i]
        else:
            for i in range(0, len(img_list1)):
                result += "["+img_list1[i]+"]"
                result += img_list2[i]
        return result


def getImg(html, choice):
    if(choice == 1):
        return getimg.youdao(html)
    if(choice == 2):
        return getimg.jinshan(html)
    if(choice == 3):
        return getimg.bing(html)
    if(choice == 4):
        return getimg.haici(html)


def url(choice):  # 选择翻译网站
    if(choice == 1):
        return "http://dict.youdao.com/w/"
    if(choice == 2):
        return "https://www.iciba.com/word?w="
    if(choice == 3):
        return "https://cn.bing.com/dict/search?q="
    if(choice == 4):
        return "https://dict.cn/search?q="


def phrase(choice, word):  # 如果是词组，就将空格替换
    if(choice == 1):
        return re.sub(' ', '%20', word)
    if(choice == 2):
        return re.sub(' ', '%20', word)
    if(choice == 3):
        return re.sub(' ', '+', word)
    if(choice == 4):
        return re.sub(' ', '+', word)


def unfind(choice):
    if (choice == 1):
        return "	中英	；"
    elif(choice == 2):
        return "未收录"
    elif(choice == 3):
        return "词典"
    elif(choice == 4):
        return "未收录"
# =======================================================================


# =======================================================================
# 图形化界面:
class Application(Frame):
    w = 500

    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.pack()
        self.createWidgets()

    def createWidgets(self):  # 图形化界面布局

        # 文件路径的选择：
        self.path_var = "仅支持txt和xlsx"
        self.xx = StringVar()
        self.result = StringVar()
        self.word = StringVar()
        Label(self, text="原文档：").grid(row=0, sticky="w")
        self.txt = Entry(self, width=70)
        self.txt.grid(row=0, column=1)
        self.txt.delete(0, "end")
        self.txt.insert(0, self.path_var)
        Label(self, text="翻译文档：").grid(row=1, sticky="w")
        self.save = Entry(self, width=70)
        self.save.grid(row=1, column=1)
        self.save.delete(0, "end")
        self.save.insert(0, self.path_var)
        Button(self, text="选择要翻\n译的文件", command=self.callback).grid(
            row=0, column=2, rowspan=2, sticky="n" + "s" + "w" + "e")

        # 词典选择栏以及翻译（运行）按钮：
        size = 4
        Label(self, text="选 择\n你 的\n词 典").grid(row=2, sticky="w", rowspan=size)
        URL = [
            ("有 道", 1),
            ("金 山", 2),
            ("bing", 3),
            ("海 词", 4)
        ]
        self.v = IntVar()
        self.v.set(1)
        for choice, num in URL:
            self.b = Radiobutton(self, text=choice, variable=self.v, value=num)
            self.b.grid(row=num+1, column=1, sticky="w")
        self.alertButton = Button(
            self, text='translate', command=self.translate)
        self.alertButton.grid(row=2, column=2, rowspan=size,
                              sticky="n" + "s" + "w" + "e")

        # 进度条：
        Label(self, text='进度:').grid(row=size+2)
        self.canvas = Canvas(self, width=self.w, height=22, bg="white")
        self.canvas.grid(row=size+2, column=1)
        Label(self, textvariable=self.xx).grid(row=size+2, column=2)

        # 每个单词的查询结果：
        Label(self, textvariable=self.word, width=8).grid(row=size+3)
        Label(self, textvariable=self.result,
              width=100).grid(row=size+3, column=1, columnspan=2)

    def change_schedule(self, now_schedule, all_schedule):
        # 进度百分比
        self.xx.set(str(round((now_schedule+1)/all_schedule*100, 2)) + '%')
        if round((now_schedule+1)/all_schedule*100, 2) == 100.00:
            self.xx.set("完成")

    def callback(self):
        # 设置可以选择的文件类型，不属于这个类型的，无法被选中
        filetypes = [('all files', '.*'),
                     ("文本文件", "*.txt"),
                     ("excel文件", "*.xlsx")]
        file_name = askopenfilename(
            title='选择单个文件', filetypes=filetypes, initialdir='./')
        self.path_var = file_name
        self.txt.delete(0, "end")
        self.txt.insert(0, self.path_var)
        self.savepath_var = re.sub('.txt', '_translated.txt', self.path_var)
        self.save.delete(0, "end")
        self.save.insert(0, self.savepath_var)

    def translate(self):
        # 翻译主程序
        try:
            title = self.txt.get()
            title_save = self.save.get()
            extension = title[-4:]

            if(extension != title_save[-4:]):
                messagebox.showerror(
                    "出问题啦", "ERROR！\n输入输出文档类型不匹配!")

            if(extension == ".txt"):
                choice = self.v.get()
                f = open(title)
                n = len(f.readlines())
                f.close()
                cnt = 0
                fill_line = self.canvas.create_rectangle(
                    1.5, 1.5, 0, 23, width=0, fill="white")
                m = 500
                x = self.w / m
                self.canvas.coords(fill_line, (0, 0, self.w, 60))
                self.master.update()
                fill_line = self.canvas.create_rectangle(
                    1.5, 1.5, 0, 23, width=0, fill="black")
                # for cnt in tqdm.trange(0, n):#控制台进度条
                for cnt in range(0, n):
                    x = x + self.w / n
                    self.canvas.coords(fill_line, (0, 0, x, 60))
                    self.master.update()
                    f = open(title)
                    w = codecs.open(title_save, "a+", "utf-8")
                    word = f.readlines()[cnt]
                    word = word.strip('\n')
                    if(len(word) > 0):
                        w.write(word+" ")
                        html = getHtml(url(choice)+phrase(choice, word))
                        ret = getImg(html, choice)
                        ret = re.sub(word, '', ret)
                        if ret == unfind(choice):
                            ret = "【!!!未查询到结果】"
                        self.word.set(word+":")
                        self.result.set(ret)
                        w.write(ret + '\n')
                        self.change_schedule(cnt, n)
                    f.close()
                    w.close()

            elif(extension == "xlsx"):
                maxofwordlen = 0
                maxofret = 0
                choice = self.v.get()
                wb = openpyxl.load_workbook(title)
                ws = wb.active
                n = ws.max_row
                cnt = 0
                fill_line = self.canvas.create_rectangle(
                    1.5, 1.5, 0, 23, width=0, fill="white")
                m = 500
                x = self.w / m
                self.canvas.coords(fill_line, (0, 0, self.w, 60))
                self.master.update()
                fill_line = self.canvas.create_rectangle(
                    1.5, 1.5, 0, 23, width=0, fill="black")
                # for cnt in tqdm.trange(0, n):#控制台进度条
                for cnt in range(n):
                    word = ws['a'+str(cnt+1)].value
                    maxofwordlen = max(maxofwordlen, len(word))
                    x = x + self.w / n
                    self.canvas.coords(fill_line, (0, 0, x, 60))
                    self.master.update()
                    html = getHtml(url(choice)+phrase(choice, word))
                    ret = getImg(html, choice)
                    ret = re.sub(word, '', ret)
                    maxofret = max(maxofret, len(ret))
                    if ret == unfind(choice):
                        if(ws['b'+str(cnt+1)].value == ""):
                            ws['b'+str(cnt+1)].value = '=IFERROR(FILTERXML(WEBSERVICE("http://fanyi.youdao.com/translate?&i="&A'+str(
                                cnt+1)+'&"&doctype=xml&version"),"//translation"),"")'
                        # 如果无法搜索到，则使用有道翻译
                    else:
                        ws['b'+str(cnt+1)].value = ret
                    self.word.set(word+":")
                    self.result.set(ret)
                    self.change_schedule(cnt, n)
                    wb.save(title)
                ws.column_dimensions["B"].width = min(70, 2*maxofret)
                ws.column_dimensions["A"].width = maxofwordlen
                wb.save(title)

            else:
                messagebox.showerror(
                    "出问题啦", "ERROR！\n不支持这种文档类型!")

        except:  # 提醒出了意外
            messagebox.showerror(
                "出问题啦", "ERROR！\n程序崩溃了\n要不要换个网址试试？")
# =======================================================================


app = Application()
choice = 0
app.master.title('单词批量查询')
app.mainloop()
